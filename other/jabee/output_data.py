import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import win32com.client
import numpy as np


def writeExcel(wfile, student):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = u'学習・教育到達目標の達成状況'

    tbl = student.getTable()

    side = Side(style='thin', color='000000')
    border = Border(top=side, left=side, bottom=side, right=side)
    alignment = Alignment(horizontal='center', vertical='center',
                          text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    fillB = PatternFill(patternType='solid', fgColor='d9d9d9')
    fillG = PatternFill(patternType='solid', fgColor='ebf1de')
    fillR = PatternFill(patternType='solid', fgColor='fde9d9')
    fillY = PatternFill(patternType='solid', fgColor='fefcd8')
    fillBlue = PatternFill(patternType='solid', fgColor='eeeeff')
    bold = Font(bold=True)

    for row in sheet:
        for cell in row:
            sheet[cell.coordinate].font = Font(name='BIZ UDPゴシック')

    sheet.cell(row=1, column=1).value = "JABEEの学習・教育到達目標の達成状況表"
    sheet.cell(row=1, column=1).font = Font(size=24, bold=True)
    sheet.merge_cells('A1:Z1')
    sheet.cell(row=1, column=1).alignment = alignment
    sheet.cell(row=1, column=1).fill = fillBlue

    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    sheet.cell(row=2, column=24).value = '2024年3月19日'

    sheet.print_area = 'A1:Z56'

    sheet.page_setup.centerHorizontally = True
    sheet.page_setup.centerVertically = True

    sheet.page_margins.left = 0.4
    sheet.page_margins.right = 0.4
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0

    sheet.cell(row=2, column=1).value = '学籍番号'
    sheet.cell(row=2, column=2).value = student.getID()
    sheet.cell(row=3, column=1).value = '氏名'
    sheet.cell(row=3, column=2).value = student.getName()
    sheet.cell(row=4, column=1).value = '生年月日'
    yy = student.getBirth()[0:4]
    mm = student.getBirth()[4:6]
    dd = student.getBirth()[6:8]

    sheet.cell(row=4, column=2).value = yy + '年' + mm + '月' + dd + '日'

    for rowIndex in range(2, 5):
        sheet.cell(row=rowIndex, column=1).font = bold
        sheet.cell(row=rowIndex, column=1).border = border
        sheet.cell(row=rowIndex, column=2).border = border
        sheet.cell(row=rowIndex, column=1).alignment = alignment
        sheet.cell(row=rowIndex, column=2).alignment = alignment

    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 10

    # table main : row [5:] #################################
    # データ入力
    # table header
    rowIndex = 6
    for zone_index in range(1, 2):
        col_ind1 = 2 + zone_index * 13  # 表の始まり = 科目名
        col_ind2 = col_ind1 + 1  # 科目群
        col_ind3 = col_ind1 + 2  # 必修選択
        col_ind4 = col_ind1 + 12  # 表の終わり
        colIndex = 0
        for colData in tbl[0]:
            cell = sheet.cell(row=rowIndex, column=col_ind1 + colIndex)
            cell.value = colData
            cell.border = border
            cell.alignment = alignment
            colIndex += 1

    rowIndex = 7
    ix = 0
    max_length = 0
    zone_index = 0
    for rowData in tbl[1:]:
        col_ind1 = 2 + zone_index * 13  # 表の始まり = 科目名
        col_ind2 = col_ind1 + 1  # 科目群
        col_ind3 = col_ind1 + 2  # 必修選択
        col_ind4 = col_ind1 + 12  # 表の終わり
        colIndex = 0
        for colData in rowData:
            cell = sheet.cell(row=rowIndex + ix, column=col_ind1 + colIndex)
            cell.value = str(colData).replace(
                "0.0", "").replace(".0", "")
            cell.border = border
            cell.alignment = alignment
            if colIndex == 0 and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

            colIndex += 1

        ix += 1
        if rowIndex + ix > 56:
            zone_index += 1
            ix = 0

    # レイアウト設定
    for zone_index in range(0, 2):
        col_ind1 = 2 + zone_index * 13  # 表の始まり = 科目名
        col_ind2 = col_ind1 + 1  # 科目群
        col_ind3 = col_ind1 + 2  # 必修選択
        col_ind4 = col_ind1 + 12  # 表の終わり

        # row 5 : 学習・教育到達目標
        rowIndex = 5
        col_ind = col_ind1 + 3
        sheet.cell(row=rowIndex, column=col_ind).value = u'学習・教育到達目標'
        col_alph1 = sheet.cell(row=rowIndex, column=col_ind).column_letter
        col_alph2 = sheet.cell(row=rowIndex, column=col_ind + 8).column_letter
        sheet.cell(row=rowIndex, column=col_ind).alignment = alignment
        sheet.cell(row=rowIndex, column=col_ind).font = bold
        sheet.cell(row=rowIndex, column=col_ind).fill = fillB
        sheet.cell(row=rowIndex, column=col_ind).border = border
        tmp_str = '{:}5:{:}5'.format(col_alph1, col_alph2)
        sheet.merge_cells(tmp_str)

        # table header
        rowIndex = 6
        for col in range(0, 12):
            cell = sheet.cell(row=rowIndex, column=col_ind1 + col)
            cell.border = border
        # 表の幅指定
        col_alph1 = sheet.cell(row=rowIndex, column=col_ind1).column_letter
        col_alph2 = sheet.cell(row=rowIndex, column=col_ind2).column_letter
        col_alph3 = sheet.cell(row=rowIndex, column=col_ind3).column_letter
        col_alph4 = sheet.cell(row=rowIndex, column=col_ind4).column_letter
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[col_alph1].width = adjusted_width
        sheet.column_dimensions[col_alph2].width = 5.5
        sheet.column_dimensions[col_alph3].width = 5.5
        for col in range(3, 12):
            cell = sheet.cell(row=rowIndex, column=col_ind1 + col)
            col_alph5 = cell.column_letter
            sheet.column_dimensions[col_alph5].width = 5.2
        sheet.column_dimensions[col_alph4].width = 1
        # table header の書式
        for col in range(0, 12):
            sheet.cell(row=rowIndex, column=col_ind1 + col).fill = fillB
            sheet.cell(row=rowIndex, column=col_ind1 + col).font = bold

    # JABEE集計表設定
    row_ind = 52
    item_name = ['項目', '要件', '習得', '不足', '判定']
    for ix in range(0, 5):  # JABEE集計表header
        cell = sheet.cell(row=row_ind + ix, column=col_ind3)
        cell.value = item_name[ix]
        cell.border = border
        cell.font = bold
        cell.fill = fillB
        cell.alignment = alignment

    if len(tbl) == 1:
        summary = np.zeros(9)
    else:
        summary = np.sum(np.array(list(zip(*tbl[1:]))[3:]), axis=1)
    for col in range(1, 10):  # JABEE集計
        col_ind = col_ind3 + col
        alph = sheet.cell(row=6, column=col_ind).column_letter

        sheet.cell(row=row_ind + 0, column=col_ind).value = col
        sheet.cell(row=row_ind + 1,
                   column=col_ind).value = student.requirement[col - 1]
        sheet.cell(row=row_ind + 2, column=col_ind).value = summary[col - 1]
        result = max(0, sheet.cell(row=row_ind + 1,
                                   column=col_ind).value - sheet.cell(row=row_ind + 2, column=col_ind).value)
        sheet.cell(row=row_ind + 3, column=col_ind).value = result
        sheet.cell(row=row_ind + 4,
                   column=col_ind).value = f'=IF({alph}55 < 1, "達成", "未達")'
        for ix in range(0, 4):
            sheet.cell(row=row_ind + ix, column=col_ind).border = border
            sheet.cell(row=row_ind + ix, column=col_ind).alignment = alignment
        sheet.cell(row=row_ind, column=col_ind).font = bold
        sheet.cell(row=row_ind, column=col_ind).fill = fillB
        sheet.cell(row=row_ind + 4, column=col_ind).font = bold
        sheet.cell(row=row_ind + 4, column=col_ind).fill = fillY

    wb.save(wfile)
    wb.close()


def excel2pdf(input_file, output_file):
    # エクセルを開く
    app = win32com.client.Dispatch("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    # Excelでワークブックを読み込む
    book = app.Workbooks.Open(input_file)
    # PDF形式で保存
    xlTypePDF = 0
    book.ExportAsFixedFormat(xlTypePDF, output_file)
    # エクセルを閉じる
    app.Quit()
